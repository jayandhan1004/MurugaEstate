from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os
import json

app = Flask(__name__)
CORS(app)

# Configuration
EXCEL_FILE = 'muruga_estates_registrations.xlsx'
ADMIN_FILE = 'admin_profile.json'

# Initialize Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        # Create headers
        headers = [
            'Full Name',
            'Contact Number',
            'Email Address',
            'Budget (Price Range)',
            'Preferred Location',
            'Villa Type',
            'Number of BHK',
            'Additional Info',
            'Registration Date & Time'
        ]
        
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 20
        
        wb.save(EXCEL_FILE)
        print(f"✓ Excel file '{EXCEL_FILE}' created successfully!")

# Initialize admin profile if it doesn't exist
def init_admin_profile():
    default_profile = {
        "ownerName": "P. Teekaraman",
        "mobile": "9962699649",
        "email": "teekaraman.ptsjn@gmail.com",
        "businessName": "Muruga Estates",
        "description": "Premium Real Estate Solutions - 50+ Successful Projects",
        "locations": "Hosur Rayakottai Road, Hosur Thally Road, Hosur Alasanatham"
    }
    
    if not os.path.exists(ADMIN_FILE):
        with open(ADMIN_FILE, 'w') as f:
            json.dump(default_profile, f, indent=4)
        print(f"✓ Admin profile file '{ADMIN_FILE}' created successfully!")
    
    return default_profile

# Route: Save Registration
@app.route('/save_registration', methods=['POST'])
def save_registration():
    try:
        data = request.get_json()
        
        # Initialize Excel if needed
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Prepare data row
        new_row = [
            data.get('fullName', ''),
            data.get('contactNumber', ''),
            data.get('emailId', ''),
            data.get('priceRange', ''),
            data.get('location', ''),
            data.get('villaType', ''),
            data.get('bhk', ''),
            data.get('additionalInfo', ''),
            data.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        # Append row to Excel
        ws.append(new_row)
        
        # Style the new row
        row_num = ws.max_row
        data_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            if row_num % 2 == 0:
                cell.fill = data_fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        
        # Save workbook
        wb.save(EXCEL_FILE)
        
        print(f"✓ New registration saved: {data.get('fullName')} - {data.get('emailId')}")
        
        return jsonify({
            'success': True,
            'message': f'Registration saved successfully! Total registrations: {row_num - 1}'
        }), 200
    
    except Exception as e:
        print(f"✗ Error saving registration: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route: Update Admin Profile
@app.route('/update_profile', methods=['POST'])
def update_profile():
    try:
        data = request.get_json()
        
        profile = {
            "ownerName": data.get('ownerName', ''),
            "mobile": data.get('mobile', ''),
            "email": data.get('email', ''),
            "businessName": data.get('businessName', ''),
            "description": data.get('description', ''),
            "locations": data.get('locations', '')
        }
        
        with open(ADMIN_FILE, 'w') as f:
            json.dump(profile, f, indent=4)
        
        print(f"✓ Admin profile updated: {profile['businessName']}")
        
        return jsonify({
            'success': True,
            'message': 'Profile updated successfully!'
        }), 200
    
    except Exception as e:
        print(f"✗ Error updating profile: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route: Get Admin Profile
@app.route('/get_profile', methods=['GET'])
def get_profile():
    try:
        if os.path.exists(ADMIN_FILE):
            with open(ADMIN_FILE, 'r') as f:
                profile = json.load(f)
        else:
            profile = init_admin_profile()
        
        return jsonify(profile), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route: Get All Registrations
@app.route('/get_registrations', methods=['GET'])
def get_registrations():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'registrations': []}), 200
        
        df = pd.read_excel(EXCEL_FILE)
        registrations = df.to_dict('records')
        
        return jsonify({
            'success': True,
            'total': len(registrations),
            'registrations': registrations
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route: Export Registrations as Excel
@app.route('/export_excel', methods=['GET'])
def export_excel():
    try:
        if os.path.exists(EXCEL_FILE):
            return {
                'success': True,
                'message': f'Excel file available: {EXCEL_FILE}',
                'file': EXCEL_FILE
            }
        else:
            return {'success': False, 'message': 'No registrations yet'}, 404
    
    except Exception as e:
        return {'error': str(e)}, 500

# Route: Download Excel File
@app.route('/download_excel', methods=['GET'])
def download_excel():
    try:
        if os.path.exists(EXCEL_FILE):
            from flask import send_file
            return send_file(EXCEL_FILE, as_attachment=True)
        else:
            return {'error': 'Excel file not found'}, 404
    
    except Exception as e:
        return {'error': str(e)}, 500

# Route: Get Statistics
@app.route('/get_statistics', methods=['GET'])
def get_statistics():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({
                'total_registrations': 0,
                'price_ranges': {},
                'locations': {},
                'villa_types': {}
            }), 200
        
        df = pd.read_excel(EXCEL_FILE)
        
        stats = {
            'total_registrations': len(df),
            'price_ranges': df['Budget (Price Range)'].value_counts().to_dict() if len(df) > 0 else {},
            'locations': df['Preferred Location'].value_counts().to_dict() if len(df) > 0 else {},
            'villa_types': df['Villa Type'].value_counts().to_dict() if len(df) > 0 else {},
            'bhk_distribution': df['Number of BHK'].value_counts().to_dict() if len(df) > 0 else {}
        }
        
        return jsonify(stats), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Health Check
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'Server is running!'}), 200

# Root Route
@app.route('/', methods=['GET'])
def home():
    return jsonify({
        'name': 'Muruga Estates API',
        'version': '1.0',
        'endpoints': {
            'POST /save_registration': 'Save client registration to Excel',
            'POST /update_profile': 'Update admin profile',
            'GET /get_profile': 'Get admin profile',
            'GET /get_registrations': 'Get all registrations',
            'GET /get_statistics': 'Get registration statistics',
            'GET /download_excel': 'Download Excel file',
            'GET /health': 'Health check'
        }
    }), 200

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🏢 MURUGA ESTATES - Real Estate Management System")
    print("="*60 + "\n")
    
    # Initialize files
    init_excel()
    init_admin_profile()
    
    print("\n" + "="*60)
    print("Starting Flask Server...")
    print("="*60)
    print("\n✓ Server running at: http://localhost:5000")
    print("✓ Excel file: muruga_estates_registrations.xlsx")
    print("✓ Admin profile: admin_profile.json\n")
    
    # Run the Flask app
    app.run(debug=True, host='localhost', port=5000)
